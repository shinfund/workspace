# -*- coding: utf-8 -*-
"""
포항영덕간 터널 시공업체 연락처 생성
- 하자보수현황 엑셀 4개(흥해/청하/남정5/점검팀) → 시공업체 연락처 엑셀 출력
- 담당자명: 날짜 기준 최신 + 이름+직함 순서 통일
- 관련업무: 하자내용/처리상세 설비명 추출 → 관리분류 fallback
- 순번: 터널별 리셋
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, OrderedDict
import datetime, re

FILES = [
    {'path': r"C:\Users\shinf\Downloads\하자보수현황_흥해터널_202606261513.xlsx",
     'division': '안전관리자', 'default_tunnel': '흥해'},
    {'path': r"C:\Users\shinf\Downloads\하자보수현황_청하터널_202606261503.xlsx",
     'division': '안전관리자', 'default_tunnel': '청하'},
    {'path': r"C:\Users\shinf\Downloads\하자보수현황_남정5터널_202606261503.xlsx",
     'division': '안전관리자', 'default_tunnel': '남정5'},
    {'path': r"C:\Users\shinf\Downloads\하자보수현황_점검팀_202606261503.xlsx",
     'division': '점검팀', 'default_tunnel': None},
]

TUNNEL_ORDER = OrderedDict([
    ('흥해', 1), ('청하', 2),
    ('남정1', 3), ('남정2', 4), ('남정3', 5), ('남정4', 6),
    ('남정5', 7), ('남정6', 8), ('남정7', 9), ('남정8', 10), ('남정9', 11),
    ('강구1', 12), ('강구2', 13), ('강구3', 14),
])
DIV_ORDER = {'안전관리자': 0, '점검팀': 1}

TITLES = [
    '대표이사', '이사장', '전무이사', '상무이사',
    '대표', '전무', '상무', '소장', '이사', '부장',
    '차장', '과장', '대리', '팀장', '실장', '수석',
    '주임', '사원',
]

def normalize_contact(s):
    s = re.sub(r'\s+', ' ', s).strip()
    for t in TITLES:
        if s.startswith(t + ' '):
            rest = s[len(t):].strip()
            if rest:
                return f"{rest} {t}"
        if s == t:
            return s
    return s

def get_tunnel_key(tunnel):
    if not tunnel:
        return (99, '')
    t = str(tunnel).strip()
    if t in TUNNEL_ORDER:
        return (TUNNEL_ORDER[t], t)
    for key, val in TUNNEL_ORDER.items():
        if key in t:
            return (val, t)
    return (98, t)

def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(c).strip() if c else '' for c in row]
        if '날짜' in vals and '관리' in vals:
            return i
    return 2

def gc(row, idx):
    if idx < 0 or idx >= len(row) or row[idx] is None:
        return ''
    return str(row[idx]).strip()

# ── 1단계: phone+company 기준 최신 contact 수집
phone_co_map = defaultdict(list)

for finfo in FILES:
    wb = openpyxl.load_workbook(finfo['path'], data_only=True)
    tgt = next((s for s in wb.sheetnames if '하자' in s), wb.sheetnames[0])
    ws = wb[tgt]
    hr = find_header_row(ws)
    hdr = [str(c).strip() if c else ''
           for c in next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))]
    def ci(name):
        for i, h in enumerate(hdr):
            if name in h: return i
        return -1
    i_date   = ci('날짜')
    i_co     = ci('업체명')
    i_person = ci('업체담당자')
    i_phone  = ci('전화번호')

    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not any(c for c in row): continue
        company = gc(row, i_co)
        if not company or company == '업체명': continue
        contact = gc(row, i_person)
        phone   = gc(row, i_phone)
        if not phone: continue
        raw_date = row[i_date] if i_date >= 0 else None
        if isinstance(raw_date, datetime.datetime):
            d = raw_date
        else:
            try:
                d = datetime.datetime.strptime(str(raw_date)[:10], '%Y-%m-%d')
            except:
                d = datetime.datetime.min
        phone_co_map[(phone, company)].append((d, contact))

def best_contact(phone, company):
    entries = phone_co_map.get((phone, company), [])
    if not entries:
        return ''
    entries_sorted = sorted(entries, key=lambda x: x[0], reverse=True)
    raw = entries_sorted[0][1]
    raw = raw.replace('점범모', '정범모')
    return normalize_contact(raw)

# ── 2단계: 연락처 집계
records = defaultdict(lambda: {'관리': set(), '텍스트': []})

for finfo in FILES:
    wb = openpyxl.load_workbook(finfo['path'], data_only=True)
    tgt = next((s for s in wb.sheetnames if '하자' in s), wb.sheetnames[0])
    ws = wb[tgt]
    hr = find_header_row(ws)
    hdr = [str(c).strip() if c else ''
           for c in next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))]
    def ci2(name):
        for i, h in enumerate(hdr):
            if name in h: return i
        return -1
    i_mgmt    = ci2('관리')
    i_tunnel  = ci2('터널명')
    i_issue   = ci2('하자내용')
    i_process = ci2('처리상세')
    i_co      = ci2('업체명')
    i_person  = ci2('업체담당자')
    i_phone   = ci2('전화번호')

    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not any(c for c in row): continue
        company = gc(row, i_co)
        if not company or company == '업체명': continue
        phone   = gc(row, i_phone)
        raw_ct  = gc(row, i_person)
        if not raw_ct and not phone: continue

        mgmt    = gc(row, i_mgmt)
        tunnel  = gc(row, i_tunnel) if i_tunnel >= 0 else ''
        tunnel  = tunnel or (finfo['default_tunnel'] or '')
        issue   = gc(row, i_issue)
        process = gc(row, i_process)

        contact = best_contact(phone, company) if phone else normalize_contact(raw_ct)

        key = (finfo['division'], tunnel, company, contact, phone)
        if mgmt:
            records[key]['관리'].add(mgmt)
        for txt in [issue, process]:
            if txt:
                records[key]['텍스트'].append(txt.replace('\n', ' ').strip()[:60])

EQUIP_MAP = [
    ('UPS', 'UPS'), ('BMS', '축전지 BMS'), ('축전지', '축전지'),
    ('밧데리', '배터리'), ('배터리', '배터리'), ('태양광', '태양광 발전기'),
    ('인버터', '인버터'), ('PRG', '자동도로전광표시기'), ('ATD', '자동도로전광표시기'),
    ('전력자동제어', '전력자동제어'), ('방재PC', '방재PC'), ('RTU', '설비RTU'),
    ('PLC', 'PLC'), ('HMI', 'HMI'), ('ACB', 'ACB'), ('VCB', 'VCB'),
    ('MCCB', 'MCCB'), ('큐비클', '큐비클'), ('배전반', '배전반'),
    ('정류기', '정류기'), ('UVR', 'UVR'), ('히팅케이블', '히팅케이블'),
    ('히팅', '히팅설비'), ('PCB기판', 'PCB기판'), ('PCB', 'PCB기판'),
    ('차단기', '차단기'), ('누전', '누전차단기'), ('X레이', 'X-Ray 회전기'),
    ('IP Wall', 'IP Wall'), ('IPWall', 'IP Wall'), ('CCTV', 'CCTV'),
    ('VICO', 'VICO'), ('비상방송', '비상방송'), ('방송', '비상방송'),
    ('비상전화', '비상전화'), ('드렌처', '드렌처'), ('스프링클러', '스프링클러'),
    ('소화전', '소화전'), ('시각경보기', '시각경보기'), ('경종', '경종'),
    ('발신기', '발신기'), ('수신기', '수신기'), ('감지기', '감지기'),
    ('DTS', '선형열감지기(DTS)'), ('자동화재', '자동화재탐지설비'), ('소화설비', '소화설비'),
    ('JET-FAN', '제트팬'), ('JETFAN', '제트팬'), ('제트팬', '제트팬'),
    ('환기', '환기설비'), ('배수펌프', '배수펌프'), ('수막', '수막설비'),
    ('배관', '배관'), ('차단막', '터널차단막'), ('하이패스', '하이패스'),
    ('누수', '누수'), ('방수', '방수'), ('균열', '균열'), ('미장', '미장'),
    ('도장', '도장'), ('외벽', '외벽'), ('바닥', '바닥'), ('공동구', '공동구'),
    ('창호', '창호'), ('화장실', '화장실'), ('조명', '조명'),
    ('콘크리트', '콘크리트'), ('사다리', '사다리'),
]

MGMT_FALLBACK = {
    '전기': '전기설비', '소방': '소방설비', '통신': '통신설비',
    '환경': '환경설비', '토목': '토목', '기계': '기계설비',
    '건축': '건축', '안전': '안전설비',
}

def extract_equip(text):
    text_up = text.upper()
    for kw, label in EQUIP_MAP:
        if kw.upper() in text_up:
            return label
    return None

def make_job(관리set, texts):
    seen, labels = set(), []
    for txt in texts:
        eq = extract_equip(txt)
        if eq and eq not in seen:
            seen.add(eq); labels.append(eq)
    if labels:
        return ', '.join(labels[:5])
    fallbacks = [MGMT_FALLBACK[m] for m in sorted(관리set) if m in MGMT_FALLBACK]
    if fallbacks:
        seen2, uniq = set(), []
        for f in fallbacks:
            if f not in seen2: seen2.add(f); uniq.append(f)
        return ', '.join(uniq)
    return '-'

# ── 3단계: 정렬
rows_data = []
for key, val in records.items():
    division, tunnel, company, contact, phone = key
    rows_data.append({
        'div': division, 'tunnel': tunnel,
        'company': company, 'contact': contact, 'phone': phone,
        'job': make_job(val['관리'], val['텍스트']),
        'note': '',
    })
rows_data.sort(key=lambda r: (
    DIV_ORDER.get(r['div'], 9),
    get_tunnel_key(r['tunnel']),
    r['company'], r['contact'],
))

# ── 4단계: 엑셀 출력
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = '시공업체 연락처'

BG = {
    'title': '1F4E79', 'hdr':   '2E75B6',
    'div_s': 'BDD7EE', 'div_i': '70AD47',
    'tun_s': 'DEEAF1', 'tun_i': 'E2EFDA',
}

def brd():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def cs(cell, bg=None, fg='000000', bold=False, ha='left', va='center',
       wrap=True, b=None, size=11):
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    else:
        cell.fill = PatternFill(fill_type=None)
    cell.font = Font(bold=bold, color=fg, size=size, name='맑은 고딕')
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if b: cell.border = b

ws_out.merge_cells('A1:G1')
tc = ws_out['A1']
tc.value = '포항영덕간 터널 시공업체 연락처'
cs(tc, bg=BG['title'], fg='FFFFFF', bold=True, ha='center', size=15, b=brd())
ws_out.row_dimensions[1].height = 36

COLS = ['순번', '관리', '터널별 업체명', '담당자', '전화번호', '관련업무', '비고']
for ci3, h in enumerate(COLS, 1):
    c = ws_out.cell(row=2, column=ci3, value=h)
    cs(c, bg=BG['hdr'], fg='FFFFFF', bold=True, ha='center', b=brd())
ws_out.row_dimensions[2].height = 22

div_sec = {'안전관리자': BG['div_s'], '점검팀': BG['div_i']}
tun_sec = {'안전관리자': BG['tun_s'], '점검팀': BG['tun_i']}
div_fg  = {'안전관리자': '1F4E79',    '점검팀': 'FFFFFF'}

prev_div, prev_tunnel = None, None
rn = 3
seq = 1

for r in rows_data:
    div, tunnel = r['div'], r['tunnel']

    if div != prev_div:
        ws_out.merge_cells(f'A{rn}:G{rn}')
        hc = ws_out.cell(row=rn, column=1, value=f'■ {div}')
        cs(hc, bg=div_sec[div], fg=div_fg[div], bold=True, ha='left', size=12, b=brd())
        ws_out.row_dimensions[rn].height = 22
        rn += 1
        prev_div, prev_tunnel, seq = div, None, 1

    if tunnel != prev_tunnel:
        ws_out.merge_cells(f'A{rn}:G{rn}')
        tc2 = ws_out.cell(row=rn, column=1, value=f'  └ {tunnel} 터널')
        cs(tc2, bg=tun_sec[div], fg='1F4E79', bold=True, ha='left', size=11, b=brd())
        ws_out.row_dimensions[rn].height = 19
        rn += 1
        prev_tunnel = tunnel
        seq = 1

    vals = [seq, div, r['company'], r['contact'], r['phone'], r['job'], r['note']]
    for ci3, v in enumerate(vals, 1):
        c = ws_out.cell(row=rn, column=ci3, value=v)
        cs(c, bg=None, ha='center' if ci3 == 1 else 'left', wrap=True, b=brd())
    ws_out.row_dimensions[rn].height = 32
    rn += 1
    seq += 1

for ci3, w in enumerate([6, 12, 22, 14, 16, 46, 14], 1):
    ws_out.column_dimensions[get_column_letter(ci3)].width = w

ws_out.freeze_panes = 'A3'

now = datetime.datetime.now().strftime('%Y%m%d%H%M')
out_path = rf"C:\Users\shinf\workspace\scripts\포항영덕간 시공업체 연락처_{now}.xlsx"
wb_out.save(out_path)
print(f"저장: {out_path}")
print(f"총 {len(rows_data)}개 연락처")
